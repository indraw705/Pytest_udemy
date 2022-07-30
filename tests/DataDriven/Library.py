import openpyxl


class Common:
    def __init__(self, file_name, sheet_name):
        global wb
        global sh
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]

    def fech_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_col_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = sh.max_column
        name_list = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            name_list.insert(i-1,cell)
        return name_list

    def update_req_with_data(self, row_num, json_req, key_list):
        c = sh.max_column
        for i in range(1, c + 1):
            cell = sh.cell(row=row_num, column=i)
            json_req[key_list[i - 1]] = cell.value

        return json_req
